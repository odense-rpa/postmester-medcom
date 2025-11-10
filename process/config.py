from typing import Dict, List, Any
from openpyxl import load_workbook

def get_excel_mapping() -> List[Dict[str, Any]]:
    """Henter excel-mapping"""
    global excel_mappings
    if not excel_mappings:
        raise ValueError("excel-mapping er ikke indlæst, brug load_excel_mapping først")
    return excel_mappings


def load_excel_mapping(file_path: str):    
    global excel_mappings
    try:
        # Load workbook and get first worksheet
        workbook = load_workbook(file_path)
        worksheet = workbook.active

        if worksheet is None:
            raise ValueError("Worksheet could not be loaded")

        # Get header row (row 1)
        header_row = worksheet[1]
        headers = []
        for cell in header_row:
            if cell.value and str(cell.value).strip():
                headers.append(str(cell.value).strip())

        # Initialize list to store row objects
        rows_list = []

        # Process each data row (starting from row 2)
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            # Create object (dictionary) for this row
            row_obj = {}
            for idx, header in enumerate(headers):
                # Get cell value at this index
                cell_value = row[idx] if idx < len(row) else None
                # Store value (convert to string and strip if not None)
                if cell_value is not None:
                    row_obj[header] = str(cell_value).strip()
                else:
                    row_obj[header] = None
            
            # Only add row if it has at least one non-None value
            if any(value is not None for value in row_obj.values()):
                rows_list.append(row_obj)

        excel_mappings = rows_list

    except Exception as e:
        raise RuntimeError(
            f"Failed to load mapping from Excel file '{file_path}': {str(e)}"
        ) from e